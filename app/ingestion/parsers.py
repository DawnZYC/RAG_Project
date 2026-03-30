"""
文档解析器 — 每种格式一个函数。
输入：文件路径
输出：list of (文本, 页码, 章节标题)
"""

import csv
from pathlib import Path
from typing import Optional

from app.models import SupportedFileType


# 类型别名，方便阅读
# 每条记录是 (文本内容, 页码, 章节标题)
ParsedPage = tuple[str, Optional[int], Optional[str]]


# ── PDF ────────────────────────────────────────────────────────────────────────

def parse_pdf(file_path: Path) -> list[ParsedPage]:
    import pdfplumber

    results: list[ParsedPage] = []
    try:
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                text = page.extract_text() or ""
                text = text.strip()
                if text:
                    results.append((text, page_num, None))
    except Exception:
        # pdfplumber 失败时用 PyMuPDF 兜底
        results = _parse_pdf_pymupdf(file_path)

    return results


def _parse_pdf_pymupdf(file_path: Path) -> list[ParsedPage]:
    import fitz  # PyMuPDF

    results: list[ParsedPage] = []
    doc = fitz.open(str(file_path))
    for page_num, page in enumerate(doc, start=1):
        text = page.get_text().strip()
        if text:
            results.append((text, page_num, None))
    doc.close()
    return results


# ── DOCX ───────────────────────────────────────────────────────────────────────

def parse_docx(file_path: Path) -> list[ParsedPage]:
    from docx import Document

    doc = Document(str(file_path))
    results: list[ParsedPage] = []
    current_section: Optional[str] = None
    buffer: list[str] = []

    for para in doc.paragraphs:
        style_name = para.style.name if para.style else ""
        text = para.text.strip()
        if not text:
            continue

        if style_name.startswith("Heading"):
            # 遇到标题：把之前积累的内容存起来，更新章节标题
            if buffer:
                results.append(("\n".join(buffer), None, current_section))
                buffer = []
            current_section = text
        else:
            buffer.append(text)

    # 把剩余内容存起来
    if buffer:
        results.append(("\n".join(buffer), None, current_section))

    return results


# ── XLSX ───────────────────────────────────────────────────────────────────────

def parse_xlsx(file_path: Path) -> list[ParsedPage]:
    import openpyxl

    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    results: list[ParsedPage] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None for cell in row):
                continue
            row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
            rows.append(row_text)

        if rows:
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            results.append((text, None, sheet_name))

    wb.close()
    return results


# ── CSV ────────────────────────────────────────────────────────────────────────

def parse_csv(file_path: Path) -> list[ParsedPage]:
    ROWS_PER_PAGE = 100  # 每 100 行打包成一页，避免文本块过大
    results: list[ParsedPage] = []

    with open(file_path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)

    for batch_idx in range(0, len(rows), ROWS_PER_PAGE):
        batch = rows[batch_idx: batch_idx + ROWS_PER_PAGE]
        text = "\n".join(" | ".join(cell for cell in row) for row in batch)
        page_num = (batch_idx // ROWS_PER_PAGE) + 1
        section = f"rows {batch_idx + 1}-{min(batch_idx + ROWS_PER_PAGE, len(rows))}"
        results.append((text, page_num, section))

    return results


# ── TXT ────────────────────────────────────────────────────────────────────────

def parse_txt(file_path: Path) -> list[ParsedPage]:
    text = file_path.read_text(encoding="utf-8", errors="replace").strip()
    return [(text, None, None)] if text else []


# ── Markdown ───────────────────────────────────────────────────────────────────

def parse_md(file_path: Path) -> list[ParsedPage]:
    import re

    text = file_path.read_text(encoding="utf-8", errors="replace")
    # 按 H1/H2 标题切分
    sections = re.split(r"(?m)^(#{1,2}\s+.+)$", text)

    results: list[ParsedPage] = []
    current_title: Optional[str] = None

    for part in sections:
        part = part.strip()
        if not part:
            continue
        if re.match(r"^#{1,2}\s+", part):
            current_title = re.sub(r"^#+\s+", "", part)
        else:
            results.append((part, None, current_title))

    return results


# ── 路由 ───────────────────────────────────────────────────────────────────────

PARSER_MAP = {
    SupportedFileType.PDF:  parse_pdf,
    SupportedFileType.DOCX: parse_docx,
    SupportedFileType.XLSX: parse_xlsx,
    SupportedFileType.CSV:  parse_csv,
    SupportedFileType.TXT:  parse_txt,
    SupportedFileType.MD:   parse_md,
}


def detect_file_type(file_path: Path) -> SupportedFileType:
    ext = file_path.suffix.lstrip(".").lower()
    try:
        return SupportedFileType(ext)
    except ValueError:
        supported = [e.value for e in SupportedFileType]
        raise ValueError(f"不支持的文件类型: .{ext}，MVP 支持: {supported}")


def parse_document(file_path: Path) -> tuple[list[ParsedPage], SupportedFileType]:
    """统一入口：自动识别文件类型并解析，返回 (解析结果, 文件类型)"""
    file_type = detect_file_type(file_path)
    parser = PARSER_MAP[file_type]
    pages = parser(file_path)
    return pages, file_type
